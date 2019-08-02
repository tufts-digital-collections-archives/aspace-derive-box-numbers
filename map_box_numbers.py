#!/usr/bin/env python3
import csv, json, sys, os
csv.field_size_limit(sys.maxsize)

from argparse import ArgumentParser, FileType
from collections import defaultdict
from getpass import getpass
from itertools import chain, islice
from types import SimpleNamespace as NS

import regex as re
import pymysql

from more_itertools import peekable, one, chunked
from openpyxl import load_workbook

from asnake.logging import setup_logging, get_logger
from asnake.aspace import ASpace
from asnake.jsonmodel import JM

def manual_mappings(filename):
    sheet = iter(one(load_workbook(os.path.expanduser(filename))))
    next(sheet) # skip headers
    return {row[0].value:row[1].value for row in sheet if row[0].value}

def omissions(filename):
    sheet = iter(one(load_workbook(os.path.expanduser(filename))))
    next(sheet) # skip headers
    return {row[0].value for row in sheet if row[0].value}

ap = ArgumentParser(description="Script to map box numbers to containers based on AO component names")
ap.add_argument('--host', default='localhost', help="host of ASpace database")
ap.add_argument('--user', default='pobocks', help='MySQL user to run as when connecting to ASpace database')
ap.add_argument('--database', default='tuftschivesspace', help="Name of MySQL database")
ap.add_argument('--omissions', type=omissions, default=set(), help="Single column Excel file with list of container barcodes to ignore")
ap.add_argument('--manual_mappings', type=manual_mappings, default={}, help='two column Excel file with mapping from barcode to indicator')
ap.add_argument('--commit', action='store_true', help='actually make changes to ASpace')
ap.add_argument('--logfile', default='map_box_numbers.log', help='path to print log to')
ap.add_argument('--cached_aos', type=FileType('r'), help='source of cached archival object jsons')
ap.add_argument('--cached_aos_save', type=FileType('w'), help='place to store cached archival object jsons')
ap.add_argument('--cached_containers', type=FileType('r'), help='source of cached container jsons')
ap.add_argument('--cached_containers_save', type=FileType('w'), help='place to store cached container jsons')

normal = re.compile(r'^(?P<coll_id>[^.]{5})\.(?P<series>\d{3})(?:\.\d{3})*\.(?P<box_no>\d{3})\.\d{5}(?:\.\d{5})?$')
box_level = re.compile(r'^(?P<coll_id>[^.]{5})\.(?P<series>\d{3})(?:\.\d{3})*\.(?P<penultimate>\d{3})\.(?P<last>\d{3})$')
weird_MS004 = re.compile(r'^(?P<coll_id>MS004)\.(?P<series>\d{3})(?:\.\d{3})*\.(?P<box_no>\d{3})\.\d{4}\.\d{2}.\d{4}$')
def sniff_box_number(component_id):
    if '-' in component_id:
        return {"box_no": "Green Barcode"}
    m = normal.match(component_id) or\
        box_level.match(component_id) or\
        weird_MS004.match(component_id)
    if m:
        return m.groupdict()

    return {"box_no": "Cannot Assign"}

def split(string, sep="."):
    return str.split(string, sep)

shared_idx = 1
def box_no_or_bust(row):
    global shared_idx
    if row['barcode'] in args.omissions:
        return 'Omitted'
    if row['barcode'] in args.manual_mappings:
        return args.manual_mappings[row['barcode']]

    indicator_prefix = ''
    if row['barcode'].endswith('g'):
        return 'Green Barcode'
    if row['shared']:
        indicator = "Shared {shared_idx}".format(shared_idx=shared_idx)
        shared_idx += 1
        return indicator
    if row['barcode'].endswith('b'):
        indicator_prefix = 'Volume '

    sniffed = [sniff_box_number(cid) for cid in  row['component_ids']]
    try:
        # boxes shared by _series_
        if all('series' in mdict for mdict in sniffed) and\
           len({mdict["series"] for mdict in sniffed}) > 1:
            coll_id = one({mdict['coll_id'] for mdict in sniffed})
            coll_name = coll_id2coll_name[coll_id]
            box_no = "{coll_name} Shared {coll_idx}".format(
                coll_name=coll_name,
                coll_idx=coll_shared_box_idxs[coll_id])
            coll_shared_box_idxs[coll_id] += 1
            return box_no


        # normal, green bc dashes, or cannot assign
        if all("box_no" in mdict for mdict in sniffed):
            box_no = one({mdict['box_no'] for mdict in sniffed})

        # Fake box level
        elif all("last" in mdict for mdict in sniffed):
            if all(mdict['last'] == '001' for mdict in sniffed):
                box_no = one({mdict['penultimate'] for mdict in sniffed})
            else:
                box_no = one({mdict['last'] for mdict in sniffed})
        else:
            return "Cannot Assign"

        return indicator_prefix + str(int(box_no))
    except ValueError:
        return "Cannot Assign"

def common_prefix(component_ids):
    pref = []
    for segments in zip(map(split, component_ids)):
        if len(set(segments)) == 1:
            pref.append(segments[0])
        else:
            break
    return pref

def convert_container_to_digital_object(container_info):
    '''Take a row representing a DGB (Digital Green Barcode) container and it's archival objects
and transform it into a digital object linked to the correct AO.'''
    # QUESTIONS:
    # - is_representative value for instance?
    # - really make a dummy file_version?
    try:
        ao_id = one(container_info['ao_ids'])
        cid = one(container_info['component_ids'])
    except Exception as e:
        import ipdb;ipdb.set_trace()
    ao = ao_jsons[ao_id]
    del ao['position'] # updating AO with position set causes issues

    digital_object = JM.digital_object(
        digital_object_id = 'tufts:{cid}'.format(cid=cid),
        title= ao['title'],
        file_versions = [JM.file_version(
            file_uri='example://no-url-available',
            publish=True
        )],
        user_defined = JM.user_defined(
            string_1 = 'Digital object location',
            text_1 = container_info['barcode']
        ),
        linked_instances = [{'ref': '/repositories/2/archival_objects/{ao_id}'.format(ao_id=ao_id)}]
    )
    log.info('create_digital_obj', digital_object=digital_object)
    if args.commit:
        d_obj_res = aspace.client.post('repositories/2/digital_objects', json=digital_object)
    else: d_obj_res = NS(status_code=200, json = lambda: {'uri': 'PLACEHOLDER'}) # mock object if dry-run
    if d_obj_res.status_code == 200:
        do_uri = d_obj_res.json()['uri']
        log.info('created_digital_object', component_id=cid, digital_object_uri=do_uri, for_real=args.commit)
        instance = JM.instance(
            instance_type='digital_object',
            digital_object={'ref': do_uri},
            is_representative=False
        )
        ao['instances'].append(instance)
        if args.commit:
            ao_res = aspace.client.post(ao['uri'], json=ao)
            if ao_res.status_code == 200:
                log.info('updated_ao', component_id=cid, ao=ao['uri'], digital_object_uri=do_uri)
                del_res = aspace.client.delete('/repositories/2/top_containers/{}'.format(container_info['container_id']))
                if del_res.status_code == 200:
                    log.info('cleanup_dgb_container', **container_info)
                else:
                    log.error('FAIL cleanup_dgb_container', result=del_res.json(), **container_info)
            else:
                log.error('FAIL updated_ao', component_id=cid, digital_object_uri=do_uri, result=ao_res.json())
                del_res = aspace.client.delete(do_uri)
                if del_res.status_code == 200:
                    log.info('digital_object_cleanup', deleted=do_uri)
                else:
                    log.error('FAIL digital_object_cleanup', deleted=do_uri, result=del_res.json())
        else:
            log.info('SKIP updated_ao', component_id=cid, message='Since this is a dry run, we shan\'t update the AO')

    else:
        log.error('FAIL created_digital_object', component_id=cid, result=d_obj_res.json())

def reindicate_container(row, new_indicator):
    '''Change indicator for container'''
    # container SHOULD be in jsons, but fallback to individual fetch if it's not for some reason?
    container = container_jsons.get(row['container_id'], None)
    if not container:
        log.warning('WARN single_container_fetch', container_id = row['container_id'])
        c_res = aspace.client.get('repositories/2/top_containers/{}'.format(row['container_id']))
        if c_res.status_code == 200:
            container = c_res.json()
            log.info('single_container_fetch', container_id = row['container_id'])
        else:
            log.error('FAIL single_container_fetch', container_id = row['container_id'])

    container['indicator'] = new_indicator
    container_res = aspace.client.post(container['uri'], json=container)
    if container_res.status_code == 200:
        log.info('updated_container', container_id=row['container_id'])
    else:
        log.info('FAIL updated_container', container_id=row['container_id'], data=row, error=container_res.json())

def map_rows(cursor):
    '''Transform JSON columns and boolean columns in resultset from MySQL'''
    for row in cursor:
        row['component_ids'] = json.loads(row['component_ids'])
        row['ao_ids'] = json.loads(row['ao_ids'])
        row['shared'] = True if row['shared'] > 1 else False
        yield row

def unmap_row(row):
    '''Transform python -> JSON for aggregate columns'''
    return {k:(json.dumps(v) if k in ('ao_ids', 'component_ids') else v) for k,v in row.items()}

def chain_aos(for_aos):
    for row in for_aos:
        yield from row['ao_ids']

if __name__ == '__main__':
    args = ap.parse_args()

    setup_logging(filename=args.logfile)
    log = get_logger('map_box_numbers')

    log.info('start')

    aspace = ASpace()
    log.info('aspace_connect')

    # note: fields match up to fields in MySQL query plus additional field for
    in_fields = ['container_id', 'barcode', 'component_ids', 'ao_ids', 'shared']
    out_fields = (*in_fields[0:2], 'proposed_box_number', *in_fields[2:],)

    conn = pymysql.connect(host=args.host, user=args.user, database=args.database, cursorclass=pymysql.cursors.DictCursor, password=
                           getpass("Please enter MySQL password for {}: ".format(args.user)))
    log.info('mysql_connect')

    with open('proposed_box_numbers.csv', 'w') as pbn,\
         open('digital_object_conversion.csv', 'w') as dgb,\
         conn.cursor() as db:

        w_pbn = csv.DictWriter(pbn, dialect='excel-tab', fieldnames=out_fields)
        w_pbn.writeheader()

        w_dgb = csv.DictWriter(dgb, dialect='excel-tab', fieldnames=in_fields)
        w_dgb.writeheader()

        shared_idx = 1

        log.info('load_coll_id2coll_name')
        db.execute('''SELECT identifier, title FROM resource''')
        coll_id2coll_name = {json.loads(row["identifier"])[0]:row["title"] for row in db.fetchall()}
        coll_shared_box_idxs = {k:1 for k in coll_id2coll_name}

        log.info('load_data')
        db.execute('''SET group_concat_max_len=995000''')
        db.execute(
            '''SELECT tc.id as container_id,
                      tc.barcode as barcode,
                      concat('[', group_concat(concat('"', ao.component_id, '"')), ']') as component_ids,
                      concat('[', group_concat(ao.id), ']') as ao_ids,
                      count(DISTINCT ao.root_record_id) as shared
               FROM top_container tc
               JOIN top_container_link_rlshp tclr
                 ON tclr.top_container_id = tc.id
               JOIN sub_container s
                 ON s.id = tclr.sub_container_id
               JOIN instance i
                 ON s.instance_id = i.id
               JOIN archival_object ao
                 ON ao.id = i.archival_object_id
               WHERE tc.indicator LIKE 'data_value_missing%'
               GROUP BY tc.indicator
               ORDER BY tc.id, tc.barcode, ao.component_id''')
        data = list(map_rows(db))
        log.info('load_data_complete')

        log.info('fetch_ao_jsons')
        ao_jsons = {}
        if args.cached_aos:
            log.info('load_aos_from_cache')
            with args.cached_aos as f:
                ao_jsons = {int(k):v for k,v in json.load(f).items()}
        else:
            for chunk in chunked(sorted(chain_aos(data)), 250):
                log.info('fetch_ao_chunk', chunk=chunk)
                ao_res = aspace.client.get('repositories/2/archival_objects', params={'id_set': chunk})
                if ao_res.status_code == 200:
                    log.info('fetch_chunk_complete', chunk="{}-{}".format(chunk[0], chunk[-1]))
                    for ao in ao_res.json():
                        ao_id = int(ao['uri'][ao['uri'].rfind('/') + 1:])
                        ao_jsons[ao_id] = ao
            if args.cached_aos_save:
                log.info('save_aos_to_cache')
                with args.cached_aos_save as f:
                    json.dump(ao_jsons, f)
        log.info('fetch_ao_jsons_complete')

        log.info('load_containers')
        if args.cached_containers:
            log.info('load_containers_from_cache')
            with args.cached_containers as f:
                container_jsons = {int(k):v for k,v in json.load(f).items()}
        else:
            container_jsons = {}
            for chunk in chunked(sorted(row['container_id'] for row in data), 250):
                log.info('fetch_container_chunk', chunk=chunk)
                c_res = aspace.client.get('repositories/2/archival_objects', params={'id_set': chunk})
                if c_res.status_code == 200:
                    log.info('fetch_chunk_complete', chunk="{}-{}".format(chunk[0], chunk[-1]))
                    for c in c_res.json():
                        c_id = int(c['uri'][c['uri'].rfind('/') + 1:])
                        container_jsons[c_id] = c
            if args.cached_containers_save:
                log.info('save_containers_to_cache')
                with args.cached_containers_save as f:
                    json.dump(container_jsons, f)

        log.info('load_containers_complete')
        log.info('data_retrieved')

        for row in data:
            if row['barcode'].startswith('DGB'):
                log.info('process_digital_barcode')
                # handle things that ought to be digital barcodes
                w_dgb.writerow(unmap_row(row))
                convert_container_to_digital_object(row)
            else:
                log.info('process_real_container')
                new_indicator = row['proposed_box_number'] = box_no_or_bust(row)

                w_pbn.writerow(unmap_row(row))
                if args.commit and new_indicator not in {'Green Barcode', 'Cannot Assign', 'Omitted'}:
                    # do the dang thing for common case
                    reindicate_container(row, new_indicator)


        log.info('end')
