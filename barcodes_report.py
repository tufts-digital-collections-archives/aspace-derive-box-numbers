#! /usr/bin/env python3
import csv, json, re

from argparse import ArgumentParser
from collections import defaultdict
from datetime import date
from getpass import getpass
from itertools import chain, repeat, groupby

import pymysql

from openpyxl import load_workbook
from more_itertools import first

from asnake.logging import setup_logging, get_logger
from asnake.aspace import ASpace
from asnake.jsonmodel import JM

ap = ArgumentParser(description="Script to determine green barcode locations")
ap.add_argument('spreadsheet', type=load_workbook, help="Spreadsheet of location barcodes")
ap.add_argument('barcode_source', type=load_workbook, help="Spreadsheet of new barcodes to be assigned")
ap.add_argument('--host', default='localhost', help="host of ASpace database")
ap.add_argument('--user', default='pobocks', help='MySQL user to run as when connecting to ASpace database')
ap.add_argument('--database', default='tuftschivesspace', help="Name of MySQL database")
ap.add_argument('--logfile', default='barcodes_report.log', help='path to print log to')

normal_component_id = re.compile(r'^(?P<coll_id>[^.]{5})\.(?P<series>\d{3})(?:\.\d{3})*\.(?P<box_no>\d{3})(?:\.\d{5}){0,2}$')

if __name__ == "__main__":
    args = ap.parse_args()
    setup_logging(filename=args.logfile)
    log = get_logger('barcodes_report')

    log.info('start')

    aspace = ASpace()
    log.info('aspace_connect')

    bc_csv_fields = [
        'original_barcode',
        'original_container_id',
        'location_id',
        'new_barcode',
        'new_container_id',
        'box_number',
        'component_id',
        'ao_id'
    ]

    loc_csv_fields = ['location_barcode', 'location_id']

    # Barcodes expected to be in first column of single-worksheet excel
    # To get the next barcode, we do: next(barcode_source)
    barcode_source = (str(first(row)) for row in args.barcode_source.worksheets[0].values)

    conn = pymysql.connect(host=args.host, user=args.user, database=args.database, cursorclass=pymysql.cursors.DictCursor,
                           password=getpass("Please enter MySQL password for {}: ".format(args.user)))
    log.info('mysql_connect')



    with open('barcode_report.csv', 'w') as barcode_report,\
         open('locations_created_report.csv', 'w') as loc_report,\
         conn:

        bc_report = csv.DictWriter(barcode_report,
                                   dialect='excel-tab',
                                   fieldnames=bc_csv_fields)
        bc_report.writeheader()
        lc_report = csv.DictWriter(loc_report,
                                   dialect='excel-tab',
                                   fieldnames=['barcode', 'location_id'])

        db = conn.cursor()
        db.execute("""SELECT barcode FROM top_container WHERE barcode REGEXP '^[0-9]+[gG]$'""")

        # Green barcodes, either from explicit list OR from matching the "digits with G as last character" format
        green_barcodes = sorted(set(chain((first(row) for row in args.spreadsheet.worksheets[0].values), (row['barcode'][0:-1] for row in db.fetchall()))))
        log.info('got_green_barcodes')

        # hash of all extant barcodes. Assumes no duplicates which is not safe in principle
        # due to lack of unique index on barcode but is safe in practice across Tufts data
        db.execute("""SELECT id, barcode FROM location WHERE barcode IS NOT NULL""")
        bc_to_loc = {row['barcode']:int(row['id']) for row in db.fetchall()}
        log.info('got_all_location_barcodes')

        missing_locations = [barcode for barcode in green_barcodes if not barcode in bc_to_loc]
        log.info('got_missing_locations', missing_locations=missing_locations)

        # hash of resource id to list of series present in resource
        db.execute("""SELECT r.id,
                             concat('["', group_concat(DISTINCT substr(ao.component_id, 7,3) SEPARATOR '","'), '"]') as series

                      FROM resource r
                      INNER JOIN archival_object ao
                        ON ao.root_record_id = r.id
                      WHERE ao.component_id REGEXP '^[A-Z]{2}[0123456789]{3}[.][0123456789]{3}[.]'
                  GROUP BY r.id""")
        rid_to_series = {str(row['id']):json.loads(row['series']) for row in db.fetchall()}
        log.info('got_resource_id_to_series')

        # Hash of f"resource_id.series" to maximum indicator in series
        db.execute('''SET group_concat_max_len=995000''')
        db.execute('''SELECT r.id,
                             substr(ao.component_id, 7, 3) as series,
                             max(CAST(regexp_substr(tc.indicator, '[0123456789]+$') AS integer)) AS max_indicator
                       FROM resource r
                       JOIN archival_object ao ON ao.root_record_id = r.id
                       JOIN instance i ON i.archival_object_id = ao.id
                       JOIN sub_container sc ON i.id = sc.instance_id
                       JOIN top_container_link_rlshp tclr ON tclr.sub_container_id = sc.id
                       JOIN top_container tc ON tc.id = tclr.top_container_id
                       WHERE tc.indicator REGEXP '[0123456789]+$'
                       AND tc.indicator REGEXP '^[0123456789;, -]+$'
                       GROUP BY r.id, series
                       HAVING max_indicator > 0
                       ORDER BY r.id''')

        series2idx = {"{}.{}".format(el['id'], el['series']):el['max_indicator'] for el in db.fetchall()}
        log.info('got_series_last_index')

        log.info('create_missing_locations')
        # create missing locations
        loc_template = JM.location(
            building='Tisch/DCA'
        )
        for loc_bc in missing_locations:
            log.info('creating_location', barcode=loc_bc)
            res = aspace.client.post('locations', json={**loc_template, 'barcode': loc_bc})
            if res.status_code == 200:
                log.info('created_location', result=res.json())
                # add newly created barcode to hash
                bc_to_loc[loc_bc] = res.json()['uri'].split('/')[-1]
                lc_report.writerow({'barcode': loc_bc, 'location_id': bc_to_loc[loc_bc]})
            else:
                log.info('FAILED_create_location', result=res.json(), status_code=res.status_code)
                lc_report.writerow({'original_barcode': loc_bc, 'location_id': 'FAILED TO CREATE'})
        # for each green barcode
        for barcode in green_barcodes:
            # going to the API for this is unexpectedly horrible, so we're cheating and going to the database
            db.execute('''SELECT ao.id, ao.component_id, tc.id AS top_container_id FROM top_container tc
                            JOIN top_container_link_rlshp tclr ON tclr.top_container_id = tc.id
                            JOIN sub_container sc ON sc.id = tclr.sub_container_id
                            JOIN instance i ON i.id = sc.instance_id
                            JOIN archival_object ao ON ao.id = i.archival_object_id
                           WHERE barcode REGEXP %s AND i.archival_object_id IS NOT NULL
                           ORDER BY ao.component_id ASC''', (fr'^{barcode}[gG]?$',))
            ao_infos = list(db.fetchall())

            if not len(ao_infos):
                log.error('empty_ao_uris', barcode=barcode)
                continue
            try:
                location_uri = f'/locations/{bc_to_loc[barcode]}'
            except KeyError as e:
                log.error('location_barcode_not_in_bc_to_loc', barcode = barcode)
                continue

            green_ao_infos = []
            normative_ao_infos = []
            for ao_info in ao_infos:
                m = normal_component_id.match(ao_info['component_id'])
                if not m:
                    green_ao_infos.append(ao_info)
                else:
                    ao_info['match'] = m
                    normative_ao_infos.append(ao_info)

            # Template for top_container object
            tc_tmpl = JM.top_container(
                type='box',
                container_locations=[JM.container_location(
                    status="current",
                    ref=location_uri,
                    start_date=date.today().isoformat()
                )]
            )
            failures = []
            top_container_uris = []
            for ao_info in green_ao_infos:
                idx = 0
                ao = aspace.client.get(f"/repositories/2/archival_objects/{ao_info['id']}").json()
                resource_id = ao['resource']['ref'].split('/')[-1]
                # If there's only one series, start numbering at last box!
                # Doing a default because at least one resource and possibly others has no normative boxes (and thus NO series)
                if len(rid_to_series.get(resource_id, [])) == 1:
                    try:
                        idx = series2idx[f"{resource_id}.{rid_to_series[resource_id][0]}"] + 1
                    except KeyError:
                        log.info('missing_series2idx', resource_id=resource_id, rid_to_series_entries=rid_to_series.get(resource_id))
                        # if we couldn't find one somehow but still had length 1, start from 1
                        series2idx[f"{resource_id}.{rid_to_series[resource_id][0]}"] = 0
                        idx = 1

                    series2idx[f"{resource_id}.{rid_to_series[resource_id][0]}"] += 1


                # Otherwise, start from scratch
                else:
                    idx += 1

                tc_json = {**tc_tmpl, "indicator": str(idx)}
                try:
                    new_barcode = next(barcode_source)
                    tc_json['barcode'] = new_barcode
                except StopIteration:
                    # If we ran out of barcodes, create one with no barcode and log the error
                    # Create placeholder in new_barcode for report
                    new_barcode = 'RAN OUT OF BARCODES'
                    log.error('ran_out_of_barcodes', old_barcode=barcode, ao_info=ao_info)

                res = aspace.client.post('repositories/2/top_containers', json=tc_json)
                if res.status_code == 200:
                    log.info('created_tc', tc=res.json(), indicator=str(idx))
                    tc_uri = res.json()['uri']
                    top_container_uris.append(tc_uri)
                    del ao['position']
                    ao['instances'].append(
                        JM.instance(
                            instance_type='mixed_materials',
                            sub_container=JM.sub_container(
                                top_container=JM.top_container(
                                    ref=tc_uri
                                )
                            )
                        )
                    )
                    ao_res = aspace.client.post(ao['uri'], json=ao)
                    if ao_res.status_code == 200:
                        log.info('ao_updated', ao=ao_res.json())
                        bc_report.writerow({'original_barcode': barcode,
                                            'original_container_id': ao_info['top_container_id'],
                                            'location_id': bc_to_loc[barcode],
                                            'new_barcode': new_barcode,
                                            'new_container_id': res.json()['id'],
                                            'box_number': str(idx),
                                            'component_id': ao_info['component_id'],
                                            'ao_id': ao_info['id']})
                    else:
                        log.info('ao_update_failed', ao=res.json(), status_code=res.status_code)
                        failures.append(ao_info)
                else:
                    log.info('create_tc_failed', tc=res.json(), status_code=res.status_code)
                    failures.append(ao_info)

            # Group AOs by series,box_no and create a top container for each box number
            def ao_infos_key(ao_info):
                return ao_info['match'].group('box_no')

            for k, group in groupby(
                    sorted(
                        normative_ao_infos,
                        key=ao_infos_key),
                    key=ao_infos_key):
                box_no = k.lstrip('0')
                aos = list(group)
                tc_json = {**tc_tmpl, "indicator": box_no}
                try:
                    new_barcode = next(barcode_source)
                    tc_json['barcode'] = new_barcode
                except StopIteration:
                    # If we ran out of barcodes, create one with no barcode and log the error
                    # Create placeholder in new_barcode for report
                    new_barcode = 'RAN OUT OF BARCODES'
                    log.error('ran_out_of_barcodes', old_barcode=barcode, aos=[x['id'] for x in aos])
                res = aspace.client.post('repositories/2/top_containers', json=tc_json)

                if res.status_code == 200:
                    log.info('created_tc', tc=res.json(), indicator=str(box_no))
                    tc_uri = res.json()['uri']
                    top_container_uris.append(tc_uri)
                    for ao_info in aos:
                        ao = aspace.client.get(f"/repositories/2/archival_objects/{ao_info['id']}").json()
                        del ao['position']
                        ao['instances'].append(
                            JM.instance(
                                instance_type='mixed_materials',
                                sub_container=JM.sub_container(
                                    top_container=JM.top_container(
                                        ref=tc_uri
                                    )
                                )
                            )
                        )
                        ao_res = aspace.client.post(ao['uri'], json=ao)
                        if ao_res.status_code == 200:
                            log.info('ao_updated', ao=ao_res.json())
                            bc_report.writerow({'original_barcode': barcode,
                                                'original_container_id': ao_info['top_container_id'],
                                                'location_id': bc_to_loc[barcode],
                                                'new_barcode': new_barcode,
                                                'new_container_id': res.json()['id'],
                                                'box_number': box_no,
                                                'component_id': ao_info['component_id'],
                                                'ao_id': ao_info['id']})
                        else:
                            log.info('ao_update_failed', ao=res.json(), status_code=res.status_code)
                            failures.append(ao_info)
                else:
                    log.info('create_tc_failed', tc=res.json(), status_code=res.status_code)
                    failures.append(ao_info)
            if not failures:
                try:
                    db.execute('SELECT id FROM top_container WHERE barcode REGEXP %s', (fr'^{barcode}[gG]?$',))
                    top_container_id = db.fetchone()['id']
                    del_res = aspace.client.delete(f'/repositories/2/top_containers/{top_container_id}')
                    if del_res.status_code == 200:
                        log.info('deleted_container', top_container_id=top_container_id)
                    else:
                        log.info('failed to delete', top_container_id=top_container_id, del_res=del_res.json(), status_code=del_res.status_code)
                except Exception as e:
                    print(e)

    log.info('end')
